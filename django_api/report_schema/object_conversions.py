import pandas as pd
import json
import openpyxl as pyxl


# def json_file_to_json_dict(json_file_path: str) -> dict:
#     """
#     Args:
#         json_file_path: File path to the JSON file
#
#     Returns:
#         Dictionary of dictionaries that represents the JSON.
#     """
#     with open(json_file_path, 'r') as json_file:
#         return json.load(json_file)
#


def json_dict_to_dataframes_dict(json_dict: dict) -> dict:
    """
    Args:
        json_dict: Dictionary of dictionaries that represents the JSON.

    Returns:
        Dictionary of Pandas dataframes where key is the sheet name of the
        dataframe while the value is the dataframe itself.
    """
    dataframes = {}
    if not isinstance(json_dict, dict):
        json_dict = json.loads(json_dict)
    for sheet_name, json_df_dict in json_dict.items():
        dataframes[sheet_name] = dict_to_dataframe(json_df_dict)

    return dataframes


def dataframes_dict_to_json_dict(dataframes_dict: dict) -> dict:
    """
    Turns a dictionary of dataframes into a dictionary of dictionary
    represntations of those dataframes.

    Args:
        dataframes_dict: Dictionary of Pandas dataframes where key is the sheet
            name of the dataframe while the value is the inner dictionary
            corresponding to the dataframe itself.

    Returns:
        Dictionary of dictionaries where key is the sheet name of the dataframe
        while the value is the dictionary representation of the dataframe.
    """
    json_dict = {}
    for sheet_name, df in dataframes_dict.items():
        json_dict[sheet_name] = dataframe_to_dict(df)

    return json_dict


def dataframes_dict_to_workbook(dataframes_dict: dict, file_path: str):
    """
    Args:
        dataframes_dict: Dictionary of Pandas dataframes where key is the sheet
            name of the dataframe while the value is the inner dictionary
            corresponding to the dataframe itself.

        file_path: File path to save the JSON file to.
            Ex: '<directory>/<inner_directory>/<file_name_without_extension>'

    Returns:
        None
    """
    wb = pyxl.Workbook()

    names = {}
    for df_name, df in dataframes_dict.items():

        if df_name[:31] in names:
            names[df_name[:31]] += 1
            sheet_name = df_name[:29] + '_' + str(names[df_name[:31]])
        else:
            names[df_name[:31]] = 1
            sheet_name = df_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        for r in pyxl.dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        ws['A1'] = df_name

    wb.remove(wb['Sheet'])
    wb.save(f'{file_path}')


def json_dict_to_json_file(json_dict: dict, file_path: str):
    """
    Args:
        json_dict: Dictionary of dictionaries where key is the sheet name of
            the dataframe while the value is the dictionary representation of
            the dataframe.

        file_path: File path to save the JSON file to.
            Ex: '<directory>/<inner_directory>/<file_name_without_extension>'

    Returns:
        None
    """
    # Dict into JSON file
    if not isinstance(json_dict, dict):
        json_dict = json.loads(json_dict)
    with open(f'{file_path}', 'w') as jsonFile:
        json.dump(json_dict, jsonFile)


def dataframe_to_dict(dataframe: object) -> dict:
    """
    Args:
        dataframe: A pandas dataframe

    Returns:
        Dictionary representation of pandas dataframe
    """
    dup_count = 1
    while True in dataframe.columns.duplicated():
        dataframe.columns = dataframe.columns.where(
            ~dataframe.columns.duplicated(), dataframe.columns + ' dp_' + str(dup_count))
        dup_count += 1

    return json.loads(dataframe.to_json(force_ascii=False))


def dict_to_dataframe(json_dict: dict) -> object:
    """
    Args:
        json_dict: A dictionary representation of Pandas dataframe

    Returns:
        Pandas dataframe construced from dictionary.
    """
    return pd.read_json(json.dumps(json_dict, ensure_ascii=False))
